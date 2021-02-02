from openpyxl import load_workbook
import time

start = time.time()

wb = load_workbook("company_delete_raw.xlsx")
ws = wb.active

col_1 = ws['I']
companys = []
for i in range(0, 2656):
    companys.append(col_1[i].value)

target_counts = 0

# i = 2
# while i <= 8412:
    # if ws.cell(row=i, column=8).value in companys:
        # ws.delete_rows(i)
        # target_counts += 1
    # if ws.cell(row=i, column=8).value is None:
        # break
    # else:
        # i += 1

for i in range(8411, 1, -1):
    if ws.cell(row=i, column=8).value in companys:
        ws.delete_rows(i)
        target_counts += 1
    if ws.cell(row=i, column=8).value is None:
        break
    else:
        pass

print(target_counts)
ws.delete_cols(9)

wb.save("delete_file.xlsx")

print(time.time()-start)
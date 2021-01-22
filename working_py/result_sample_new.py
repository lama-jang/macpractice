from openpyxl import load_workbook

wb = load_workbook("result_sample_new.xlsx")
ws = wb.active


contents = [] # 부적합 내용 전체 리스트


def delete_insert_cols(fact_all, fact_1, fact_2):
    ws.delete_cols(12+fact_1)
    ws.insert_cols(12+fact_all)
    ws.cell(4, 12+fact_all, "부적합 내용")


# 특정 문자(O)를 찾아서 지우는 코드 필요 

# 부적합 요소 수를 함수에 설정
def result_fault(fact_all):
    
    for row in ws.iter_rows(min_row=5, min_col=12, max_col=11+fact_all):
        # 특정 문자(O)를 찾아서 지우는 코드 필요 
        
        content = [] # 1개 제품에 대한 부적합내용 리스트
        for i in range(0, fact_all): # 부적합 요소 수(12개)
            if row[i].value is not None:
                content.append("{} : {}".format(ws.cell(row=4, column=i+fact_all).value, row[i].value))
        contents.append(", ".join(content))


    # 부적합내용 리스트를 해당 시료 셀에 분배하여 입력
    for i in range(0, len(contents)):
        ws.cell(i+5, 12+fact_all, contents[i])

fact_1 = input("안전요건 부적합 요소 수 : ")
fact_2 = input("표시사항 부적합 요소 수 : ")
fact_all = int(fact_1) + int(fact_2)

delete_insert_cols(int(fact_all), int(fact_1), int(fact_2))
result_fault(int(fact_all))

wb.save("result_sample_ver2.xlsx")
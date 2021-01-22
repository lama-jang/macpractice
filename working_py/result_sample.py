from openpyxl import load_workbook

wb = load_workbook("result_sample2.xlsx")
ws = wb.active

# 안전요건 !!
contents_1 = [] # 안전요건 부적합내용 전체 리스트

for row in ws.iter_rows(min_row=5, min_col=12, max_col=14):
    content_1 = [] # 1개 제품에 대한 안전요건 부적합내용 리스트
    for i in range(0, 3): # 안전요건 부적합 요소 수(3개)
        if row[i].value is not None:
            content_1.append("{} : {}".format(ws.cell(row=4, column=i+12).value, row[i].value))
    contents_1.append(", ".join(content_1))

# 안전요건 부적합내용 리스트를 해당 시료 셀에 분배하여 입력
for i in range(0, len(contents_1)):
    ws.cell(i+5, 15, contents_1[i])


# 표시사항 !!
contents_2 = [] # 표시사항 부적합내용 전체 리스트

for row in ws.iter_rows(min_row=5, min_col=17, max_col=25):
    content_2 = [] # 1개 제품에 대한 표시사항 부적합내용 리스트
    for i in range(0, 9): # 표시사항 부적합 요소 수(9개)
        if row[i].value is not None:
            content_2.append("{} : {}".format(ws.cell(row=4, column=i+17).value, row[i].value))
    contents_2.append(", ".join(content_2))

# 안전요건 부적합내용 리스트를 해당 시료 셀에 분배하여 입력
for i in range(0, len(contents_2)):
    ws.cell(i+5, 26, contents_2[i])

wb.save("result_sample_ver1.xlsx")
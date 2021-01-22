from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 시트를 기본이름으로 생성
ws.title = "My Sheet" # 시트이름이 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB형태로 값을 입력해주면 탭 생상 변경

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 인덱스에 시트 생성

new_ws = wb["NewSheet"] # Dict 형태로 시트이름으로 접근 가능

# 시트 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

print(wb.sheetnames) # 모든 시트 이름 확인

wb.save("sample.xlsx")

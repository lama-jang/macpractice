from openpyxl import load_workbook

wb = load_workbook("test_company.xlsx")
ws = wb.active

# a+1 번째 업체명 출력
def name_company(a):
    for col in ws.iter_cols(min_row=2, max_row=3, min_col=3, max_col=3):
        return col[a].value

print(name_company(0))

# names 리스트로 업체명 모으기
names = []
for i in range(0, 2):
    names.append(name_company(i))

print(names)




name_company = []
for col in ws.iter_cols(min_row=2, max_row=3, min_col=3, max_col=3):
    for cell in col:
        name_company.append(cell.value)

model_company = []
for col in ws.iter_cols(min_row=2, max_row=3, min_col=2, max_col=2):
    for cell in col:
        model_company.append(cell.value)

address_company = []
for col in ws.iter_cols(min_row=2, max_row=3, min_col=4, max_col=4):
    for cell in col:
        address_company.append(cell.value)

number_company = []
for col in ws.iter_cols(min_row=2, max_row=3, min_col=5, max_col=5):
    for cell in col:
        number_company.append(cell.value)

cause_company = []
for col in ws.iter_cols(min_row=2, max_row=3, min_col=6, max_col=6):
    for cell in col:
        cause_company.append(cell.value)

#print(name_company)
#print(model_company)
#print(address_company)
#print(number_company)
#print(cause_company)
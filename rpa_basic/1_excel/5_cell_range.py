from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기 
ws.append(["번호", "영어", "수학"]) # A, B, C column
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 colmn만 가져오기
print(col_B)
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # 영어, 수학 column 함께 가져오기
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1] # 1번째 row만 가져오기
for cell in row_title:
    print(cell.value)

#row_range = ws[2:6] # 1번째 row 제외하고 2~6 까지 값 가져오기
#for rows in row_range:
#    for cell in rows:
#        print(cell.value, end=" ")
#    print()

from openpyxl.utils.cell import coordinate_from_string

row_range_2 = ws[2:ws.max_row]
for rows in row_range_2:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ") # 샐의 좌표정보 가져오기
        xy = coordinate_from_string(cell.coordinate) # A/1, AZ/100 튜플 형태로 row, column 좌표 정보 끊어주기
        print(xy[0], end="") # A
        print(xy[1], end=" ") # 1
    print()

# 전체 rows
# print(tuple(ws.rows)) # 한 row씩 tuple로 가져옴
# for row in tuple(ws.rows):
#     print(row[1].value) # 1번째 row 값 가져옴

# 전체 columns
# print(tuple(ws.columns)) # 한 column씩 tuple로 가져옴
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): # 전체 row
#    print(row[1].value)

# for column in ws.iter_cols():
#     print(column[0].value)

# 2번째 줄부터 11번째 줄, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3): # 좌우좌우로 1줄씩 데이터가져오기
    print(row[0].value, row[1].value) # 영어, 수학

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3): # 상하상하로 1열씩 데이터가져오기
    print(col) 


wb.save("sample.xlsx")